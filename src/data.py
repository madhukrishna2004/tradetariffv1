import json
import openpyxl

def clean_tariff_rate(value):
    if isinstance(value, (float, int)):
        value = f"{round(value * 100, 2)}%"  # Assuming the value is a percentage
    return value

def excel_to_json(input_file: str, output_file: str, sheet_name: str = "Sheet1"):
    workbook = openpyxl.load_workbook(input_file, data_only=True)
    
    # List all sheet names and print them for debugging
    sheet_names = workbook.sheetnames
    print(f"Sheet names in the workbook: {sheet_names}")
    
    if sheet_name not in sheet_names:
        raise ValueError(f"No sheet named '{sheet_name}' found in the workbook.")
    
    worksheet = workbook[sheet_name]
    data = []
    
    # Extract the header
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    
    # Iterate through rows, skipping the header
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        # Skip rows with essential missing data
        if not row[0] or not row[1]:  # Assuming commodity and description are in the first two columns
            print(f"Row {row_number}: Missing essential data (commodity or description).")
            continue
        
        # Extract row values
        data_row = {
            "commodity": str(row[0]).strip() if row[0] else None,
            "description": str(row[1]).strip() if row[1] else None,
            "cet_duty_rate": clean_tariff_rate(row[2]) if len(row) > 2 and row[2] else None,
            "ukgt_duty_rate": clean_tariff_rate(row[3]) if len(row) > 3 and row[3] else None,
            "change": str(row[4]).strip() if len(row) > 4 and row[4] else None,
            "trade_remedy_applies": str(row[5]).strip() if len(row) > 5 and row[5] else None,
            "cet_applies_until_trade_remedy_transition_reviews_concluded": str(row[6]).strip() if len(row) > 6 and row[6] else None,
            "suspension_applies": str(row[7]).strip() if len(row) > 7 and row[7] else None,
            "atq_applies": str(row[8]).strip() if len(row) > 8 and row[8] else None,
            "product_specific_rule_of_origin": str(row[9]).strip() if len(row) > 9 and row[9] else None,
            "vat_rate": clean_tariff_rate(row[10]) if len(row) > 10 and row[10] else None,  # Add VAT Rate to JSON structure
            "product_specific_rule_of_origin_japan": str(row[11]).strip() if len(row) > 11 and row[11] else None,
        }
        data.append(data_row)

    # Write to JSON
    with open(output_file, "w") as f:
        json.dump(data, f, indent=2)

if __name__ == "__main__":
    input_file = r"D:\ai project\global-uk-tariffv1\global-uk-tariff.xlsx"
    output_file = r"D:\ai project\global-uk-tariffv1\data.json"
    excel_to_json(input_file, output_file)
